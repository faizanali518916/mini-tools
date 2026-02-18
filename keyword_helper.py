#!/usr/bin/env python3
"""
Keyword Helper - Interactive CLI tool for managing keyword Excel sheets.
"""

import os
import sys
import copy
from pathlib import Path

import questionary
from questionary import Style
from rich.console import Console
from rich.table import Table
from rich.panel import Panel
from rich.text import Text
from rich import box
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

# ── Rich console ──────────────────────────────────────────────────────────────
console = Console()

# ── Questionary custom style ─────────────────────────────────────────────────
custom_style = Style(
    [
        ("qmark", "fg:#E91E63 bold"),  # pink question mark
        ("question", "fg:#FFFFFF bold"),  # white question text
        ("answer", "fg:#00BCD4 bold"),  # cyan answer
        ("pointer", "fg:#E91E63 bold"),  # pink pointer arrow
        ("highlighted", "fg:#E91E63 bold"),  # pink highlighted choice
        ("selected", "fg:#00BCD4"),  # cyan selected
        ("separator", "fg:#757575"),  # grey separator
        ("instruction", "fg:#9E9E9E"),  # grey instruction
        ("text", "fg:#FFFFFF"),  # white text
    ]
)

# ── Fill styles for Excel ────────────────────────────────────────────────────
RED_FILL = PatternFill(start_color="FF4C4C", end_color="FF4C4C", fill_type="solid")
GREEN_FILL = PatternFill(start_color="4CFF4C", end_color="4CFF4C", fill_type="solid")
WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

# ── Paths ────────────────────────────────────────────────────────────────────
BASE_DIR = Path(__file__).resolve().parent
INPUT_DIR = BASE_DIR / "input"
OUTPUT_DIR = BASE_DIR / "output"


# ══════════════════════════════════════════════════════════════════════════════
#  Helpers
# ══════════════════════════════════════════════════════════════════════════════


def clear():
    os.system("cls" if os.name == "nt" else "clear")


def banner():
    """Print the application banner."""
    clear()
    console.print(
        Panel(
            Text("⌨  KEYWORD HELPER", justify="center", style="bold white"),
            subtitle="[dim]Interactive keyword management tool[/dim]",
            border_style="bright_magenta",
            box=box.DOUBLE_EDGE,
            padding=(1, 4),
        )
    )
    console.print()


def get_row_color(row):
    """Return 'red', 'green', or None based on the fill of the first cell."""
    cell = row[0]
    fill = cell.fill
    if fill and fill.start_color and fill.start_color.rgb:
        rgb = str(fill.start_color.rgb)
        if rgb in ("00FF4C4C", "FF4C4C", "FFFF4C4C"):
            return "red"
        if rgb in ("004CFF4C", "4CFF4C", "FF4CFF4C"):
            return "green"
    return None


def get_data_rows(ws):
    """Return list of row indices that have data in the first column (skip empty rows)."""
    data_rows = []
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=1)
        if cell.value is not None:
            data_rows.append(row_idx)
    return data_rows


def sort_sheet(ws):
    """
    Sort sheet in-place so green rows are on top, uncolored in the middle,
    and red rows at the bottom.  Row 1 is treated as a header and kept.
    """
    data_rows = get_data_rows(ws)
    if not data_rows:
        return

    # Collect all data rows with their styles
    rows_data = []
    for row_idx in data_rows:
        row_cells = []
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            row_cells.append(
                {
                    "value": cell.value,
                    "fill": copy.copy(cell.fill),
                    "font": copy.copy(cell.font),
                }
            )
        rows_data.append((row_idx, row_cells))

    # Determine color for sorting
    def sort_key(item):
        row_idx, row = item
        fill = row[0]["fill"]
        if fill and fill.start_color and fill.start_color.rgb:
            rgb = str(fill.start_color.rgb)
            if rgb in ("004CFF4C", "4CFF4C", "FF4CFF4C"):
                return 0  # green first
            if rgb in ("00FF4C4C", "FF4C4C", "FFFF4C4C"):
                return 2  # red last
        return 1  # uncolored middle

    rows_data.sort(key=sort_key)

    # Write back to original positions
    for i, (orig_row_idx, row) in enumerate(rows_data):
        dest_row_idx = data_rows[i]
        for col_idx, cell_data in enumerate(row, start=1):
            cell = ws.cell(row=dest_row_idx, column=col_idx)
            cell.value = cell_data["value"]
            cell.fill = cell_data["fill"]
            cell.font = cell_data["font"]


# ══════════════════════════════════════════════════════════════════════════════
#  Core workflow
# ══════════════════════════════════════════════════════════════════════════════


def pick_file():
    """Let the user pick an .xlsx file from input/."""
    banner()
    xlsx_files = sorted(INPUT_DIR.glob("*.xlsx"))
    if not xlsx_files:
        console.print(
            "[bold red]  No .xlsx files found in input/ directory.[/bold red]"
        )
        console.print("[dim]  Place your Excel files there and try again.[/dim]\n")
        sys.exit(1)

    choices = [f.name for f in xlsx_files]
    chosen = questionary.select(
        "Select an Excel file:",
        choices=choices,
        style=custom_style,
        instruction="(↑/↓ to move, Enter to select)",
    ).ask()

    if chosen is None:
        sys.exit(0)

    return INPUT_DIR / chosen


def pick_sheet(wb):
    """Let the user pick one or multiple worksheets."""
    banner()
    sheets = wb.sheetnames

    # Ask if user wants single or multiple sheets
    mode = questionary.select(
        "How many sheets would you like to work with?",
        choices=["Single sheet", "Multiple sheets (merged)"],
        style=custom_style,
    ).ask()

    if mode is None:
        sys.exit(0)

    if "Single" in mode:
        chosen = questionary.select(
            "Select a worksheet:",
            choices=sheets,
            style=custom_style,
            instruction="(↑/↓ to move, Enter to select)",
        ).ask()
        if chosen is None:
            sys.exit(0)
        return [chosen]
    else:
        chosen = questionary.checkbox(
            "Select worksheets to merge:",
            choices=sheets,
            style=custom_style,
            instruction="(↑/↓ to move, Space to select, Enter to confirm)",
        ).ask()
        if not chosen:
            sys.exit(0)
        return chosen


def view_data(ws):
    """Display all keywords (first column) in a Rich table with pagination if needed."""
    banner()
    data_rows = get_data_rows(ws)
    total = len(data_rows)
    page_size = 500
    total_pages = (total + page_size - 1) // page_size

    # If more than one page, let user choose
    if total_pages > 1:
        page_choices = [
            f"Page {i+1}  ({i*page_size + 1}-{min((i+1)*page_size, total)})"
            for i in range(total_pages)
        ]
        banner()
        page_choice = questionary.select(
            f"Select page to view ({total_pages} pages total):",
            choices=page_choices,
            style=custom_style,
        ).ask()
        if page_choice is None:
            return
        page_num = int(page_choice.split()[1]) - 1
    else:
        page_num = 0

    start_idx = page_num * page_size
    end_idx = min(start_idx + page_size, total)
    page_data = data_rows[start_idx:end_idx]

    banner()
    page_info = f" [Page {page_num + 1}/{total_pages}]" if total_pages > 1 else ""
    table = Table(
        title=f"📋  Keywords in '{ws.title}'{page_info}",
        box=box.ROUNDED,
        border_style="bright_magenta",
        header_style="bold bright_cyan",
        show_lines=False,
        padding=(0, 2),
    )
    table.add_column("#", style="dim", width=6, justify="right")
    table.add_column("Keyword", style="white", min_width=30)
    table.add_column("Status", justify="center", width=12)

    for display_idx, row_idx in enumerate(page_data, start=start_idx + 1):
        row = [ws.cell(row=row_idx, column=1)]
        cell = row[0]
        keyword = str(cell.value) if cell.value is not None else ""
        color = get_row_color(row)
        if color == "green":
            status = "[bold green]✔ Relevant[/bold green]"
            kw_style = "green"
        elif color == "red":
            status = "[bold red]✘ Irrelevant[/bold red]"
            kw_style = "red"
        else:
            status = "[dim]—[/dim]"
            kw_style = "white"
        table.add_row(str(display_idx), f"[{kw_style}]{keyword}[/{kw_style}]", status)

    console.print(table)
    console.print()

    # Stats
    green_count = sum(
        1 for idx in data_rows if get_row_color([ws.cell(row=idx, column=1)]) == "green"
    )
    red_count = sum(
        1 for idx in data_rows if get_row_color([ws.cell(row=idx, column=1)]) == "red"
    )
    unmarked = total - green_count - red_count

    stats = Table(box=box.SIMPLE, show_header=False, padding=(0, 1))
    stats.add_column(style="bold")
    stats.add_column()
    stats.add_row("[bright_cyan]Total[/bright_cyan]", str(total))
    stats.add_row("[green]Relevant[/green]", str(green_count))
    stats.add_row("[red]Irrelevant[/red]", str(red_count))
    stats.add_row("[dim]Unmarked[/dim]", str(unmarked))
    console.print(
        Panel(stats, title="Statistics", border_style="bright_magenta", box=box.ROUNDED)
    )
    console.print()

    questionary.press_any_key_to_continue(style=custom_style).ask()


def keyword_search(ws, wb, filepath):
    """
    Search for a substring in the first column (case-insensitive).
    Then let the user mark matches as relevant or irrelevant.
    """
    banner()
    query = questionary.text(
        "Enter search keyword / substring:",
        style=custom_style,
    ).ask()

    if query is None:
        return

    query_lower = query.strip().lower()
    if not query_lower:
        console.print("[yellow]  Empty search query. Returning to menu.[/yellow]\n")
        questionary.press_any_key_to_continue(style=custom_style).ask()
        return

    # Find matching rows
    data_rows = get_data_rows(ws)
    matches = []
    for row_idx in data_rows:
        cell = ws.cell(row=row_idx, column=1)
        value = str(cell.value).lower() if cell.value is not None else ""
        if query_lower in value:
            matches.append(row_idx)

    banner()
    if not matches:
        console.print(
            f"[yellow]  No keywords found containing '[bold]{query}[/bold]'.[/yellow]\n"
        )
        questionary.press_any_key_to_continue(style=custom_style).ask()
        return

    # Show matches
    table = Table(
        title=f"🔍  Results for '{query}'  ({len(matches)} matches)",
        box=box.ROUNDED,
        border_style="bright_cyan",
        header_style="bold bright_cyan",
        show_lines=False,
        padding=(0, 2),
    )
    table.add_column("#", style="dim", width=6, justify="right")
    table.add_column("Row", style="dim", width=6, justify="right")
    table.add_column("Keyword", style="white", min_width=30)

    for i, row_idx in enumerate(matches, start=1):
        keyword = str(ws.cell(row=row_idx, column=1).value or "")
        # Highlight the matched portion
        lower_kw = keyword.lower()
        start = lower_kw.find(query_lower)
        if start != -1:
            before = keyword[:start]
            match = keyword[start : start + len(query_lower)]
            after = keyword[start + len(query_lower) :]
            display = (
                f"{before}[bold bright_magenta]{match}[/bold bright_magenta]{after}"
            )
        else:
            display = keyword
        table.add_row(str(i), str(row_idx), display)

    console.print(table)
    console.print()

    # Ask what to do with the matches
    action = questionary.select(
        "What would you like to do with these results?",
        choices=[
            "🟢  Mark all as RELEVANT (green)",
            "🔴  Mark all as IRRELEVANT (red)",
            "⊘  Mark all as NEUTRAL (unmarked)",
            "↩  Back to menu (no changes)",
        ],
        style=custom_style,
    ).ask()

    if action is None or "Back" in action:
        return

    if "NEUTRAL" in action:
        fill = WHITE_FILL
        label = "neutral"
        target_color = None
    elif "IRRELEVANT" in action:
        fill = RED_FILL
        label = "irrelevant"
        target_color = "red"
    else:
        fill = GREEN_FILL
        label = "relevant"
        target_color = "green"

    # Apply color to matched rows (with protection logic)
    marked_count = 0
    skipped_count = 0
    for row_idx in matches:
        row = [ws.cell(row=row_idx, column=1)]
        current_color = get_row_color(row)

        # Protection logic: if trying to mark as irrelevant but already relevant, skip
        # Or if trying to mark as relevant but already irrelevant, skip
        if target_color == "red" and current_color == "green":
            skipped_count += 1
            continue
        elif target_color == "green" and current_color == "red":
            skipped_count += 1
            continue

        # Apply color to all columns
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row=row_idx, column=col_idx).fill = fill
        marked_count += 1

    # Sort the sheet
    sort_sheet(ws)

    # Save
    wb.save(filepath)

    if skipped_count > 0:
        console.print(
            f"\n[bold green]  ✔ Marked {marked_count} rows as {label} and re-sorted the sheet.\n"
            f"  ⊘ Skipped {skipped_count} rows (protected by existing marking)[/bold green]\n"
        )
    else:
        console.print(
            f"\n[bold green]  ✔ Marked {marked_count} rows as {label} and re-sorted the sheet.[/bold green]\n"
        )
    questionary.press_any_key_to_continue(style=custom_style).ask()


def export_clean_sheet(ws, wb, filepath):
    """
    Export data to a new worksheet, excluding irrelevant (red) rows.
    """
    banner()

    # Suggest a name
    base_name = ws.title
    new_name = f"{base_name}_clean"
    counter = 1
    while new_name in wb.sheetnames:
        counter += 1
        new_name = f"{base_name}_clean_{counter}"

    name = questionary.text(
        "Name for the new clean sheet:",
        default=new_name,
        style=custom_style,
    ).ask()

    if name is None:
        return

    name = name.strip() or new_name

    # Create new sheet
    new_ws = wb.create_sheet(title=name)

    # Copy header
    for col_idx in range(1, ws.max_column + 1):
        src = ws.cell(row=1, column=col_idx)
        dst = new_ws.cell(row=1, column=col_idx)
        dst.value = src.value
        dst.fill = copy.copy(src.fill)
        dst.font = copy.copy(src.font)

    # Copy non-red rows from data rows only
    data_rows = get_data_rows(ws)
    dest_row = 2
    skipped = 0
    for row_idx in data_rows:
        row_cells = [
            ws.cell(row=row_idx, column=c) for c in range(1, ws.max_column + 1)
        ]
        color = get_row_color(row_cells)
        if color == "red":
            skipped += 1
            continue
        for col_idx in range(1, ws.max_column + 1):
            src = ws.cell(row=row_idx, column=col_idx)
            dst = new_ws.cell(row=dest_row, column=col_idx)
            dst.value = src.value
            dst.fill = copy.copy(src.fill)
            dst.font = copy.copy(src.font)
        dest_row += 1

    wb.save(filepath)

    total_copied = dest_row - 2
    console.print(
        Panel(
            f"[bold green]✔  Created sheet '[bright_cyan]{name}[/bright_cyan]'\n\n"
            f"   Rows copied:   [white]{total_copied}[/white]\n"
            f"   Rows removed:  [red]{skipped}[/red] (irrelevant)[/bold green]",
            border_style="green",
            box=box.ROUNDED,
            padding=(1, 3),
        )
    )
    console.print()
    questionary.press_any_key_to_continue(style=custom_style).ask()


def deduplicate_keywords(ws, wb, filepath):
    """Remove duplicate keywords (case-insensitive), keeping the first occurrence."""
    banner()

    data_rows = get_data_rows(ws)
    seen = set()
    rows_to_delete = []

    for row_idx in data_rows:
        val = ws.cell(row=row_idx, column=1).value
        key = str(val).strip().lower() if val is not None else ""
        if key in seen:
            rows_to_delete.append(row_idx)
        else:
            seen.add(key)

    if not rows_to_delete:
        console.print(
            "[bold green]  ✔ No duplicate keywords found. Sheet is already unique.[/bold green]\n"
        )
        questionary.press_any_key_to_continue(style=custom_style).ask()
        return

    confirm = questionary.confirm(
        f"Found {len(rows_to_delete)} duplicate rows. Remove them?",
        default=True,
        style=custom_style,
    ).ask()

    if not confirm:
        return

    # Delete from bottom up so row indices don't shift
    for row_idx in reversed(rows_to_delete):
        ws.delete_rows(row_idx, 1)

    wb.save(filepath)

    new_data_rows = get_data_rows(ws)
    console.print(
        f"\n[bold green]  ✔ Removed {len(rows_to_delete)} duplicate rows. "
        f"{len(new_data_rows)} unique keywords remain.[/bold green]\n"
    )
    questionary.press_any_key_to_continue(style=custom_style).ask()


# ══════════════════════════════════════════════════════════════════════════════
#  Multi-sheet operations
# ══════════════════════════════════════════════════════════════════════════════


def get_multi_sheet_data(wb, sheet_names):
    """Merge data from multiple sheets. Returns dict mapping sheet_name -> list of (row_idx, keyword)."""
    merged_data = {}
    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        data_rows = get_data_rows(ws)
        merged_data[sheet_name] = [
            (row_idx, str(ws.cell(row=row_idx, column=1).value or ""))
            for row_idx in data_rows
        ]
    return merged_data


def view_multi_sheets(wb, sheet_names):
    """Display merged keywords from multiple sheets with pagination if needed."""
    banner()
    merged_data = get_multi_sheet_data(wb, sheet_names)

    # Flatten all rows for pagination
    all_rows = []
    for sheet_name in sheet_names:
        for row_idx, keyword in merged_data[sheet_name]:
            all_rows.append((sheet_name, row_idx, keyword))

    total = len(all_rows)
    page_size = 500
    total_pages = (total + page_size - 1) // page_size

    # If more than one page, let user choose
    if total_pages > 1:
        page_choices = [
            f"Page {i+1}  ({i*page_size + 1}-{min((i+1)*page_size, total)})"
            for i in range(total_pages)
        ]
        banner()
        page_choice = questionary.select(
            f"Select page to view ({total_pages} pages total):",
            choices=page_choices,
            style=custom_style,
        ).ask()
        if page_choice is None:
            return
        page_num = int(page_choice.split()[1]) - 1
    else:
        page_num = 0

    start_idx = page_num * page_size
    end_idx = min(start_idx + page_size, total)
    page_rows = all_rows[start_idx:end_idx]

    banner()
    page_info = f" [Page {page_num + 1}/{total_pages}]" if total_pages > 1 else ""
    table = Table(
        title=f"📋  Merged Keywords from {len(sheet_names)} sheets{page_info}",
        box=box.ROUNDED,
        border_style="bright_magenta",
        header_style="bold bright_cyan",
        show_lines=False,
        padding=(0, 2),
    )
    table.add_column("#", style="dim", width=6, justify="right")
    table.add_column("Sheet", style="bright_magenta", width=15)
    table.add_column("Keyword", style="white", min_width=30)
    table.add_column("Status", justify="center", width=12)

    for display_idx, (sheet_name, row_idx, keyword) in enumerate(
        page_rows, start=start_idx + 1
    ):
        ws = wb[sheet_name]
        row = [ws.cell(row=row_idx, column=1)]
        color = get_row_color(row)
        if color == "green":
            status = "[bold green]✔ Relevant[/bold green]"
            kw_style = "green"
        elif color == "red":
            status = "[bold red]✘ Irrelevant[/bold red]"
            kw_style = "red"
        else:
            status = "[dim]—[/dim]"
            kw_style = "white"
        table.add_row(
            str(display_idx), sheet_name, f"[{kw_style}]{keyword}[/{kw_style}]", status
        )

    console.print(table)
    console.print()

    # Stats
    green_count = 0
    red_count = 0
    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        for row_idx, _ in merged_data[sheet_name]:
            row = [ws.cell(row=row_idx, column=1)]
            if get_row_color(row) == "green":
                green_count += 1
            elif get_row_color(row) == "red":
                red_count += 1
    unmarked = total - green_count - red_count

    stats = Table(box=box.SIMPLE, show_header=False, padding=(0, 1))
    stats.add_column(style="bold")
    stats.add_column()
    stats.add_row("[bright_cyan]Total[/bright_cyan]", str(total))
    stats.add_row("[green]Relevant[/green]", str(green_count))
    stats.add_row("[red]Irrelevant[/red]", str(red_count))
    stats.add_row("[dim]Unmarked[/dim]", str(unmarked))
    console.print(
        Panel(stats, title="Statistics", border_style="bright_magenta", box=box.ROUNDED)
    )
    console.print()

    questionary.press_any_key_to_continue(style=custom_style).ask()


def keyword_search_multi(wb, sheet_names, filepath):
    """Search across multiple sheets and mark matches."""
    banner()
    query = questionary.text(
        "Enter search keyword / substring:",
        style=custom_style,
    ).ask()

    if query is None:
        return

    query_lower = query.strip().lower()
    if not query_lower:
        console.print("[yellow]  Empty search query. Returning to menu.[/yellow]\n")
        questionary.press_any_key_to_continue(style=custom_style).ask()
        return

    # Find matching rows across all sheets
    matches = {}  # sheet_name -> list of row indices
    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        data_rows = get_data_rows(ws)
        sheet_matches = []
        for row_idx in data_rows:
            cell = ws.cell(row=row_idx, column=1)
            value = str(cell.value).lower() if cell.value is not None else ""
            if query_lower in value:
                sheet_matches.append(row_idx)
        if sheet_matches:
            matches[sheet_name] = sheet_matches

    banner()
    if not matches:
        console.print(
            f"[yellow]  No keywords found containing '[bold]{query}[/bold]' in selected sheets.[/yellow]\n"
        )
        questionary.press_any_key_to_continue(style=custom_style).ask()
        return

    # Show matches
    total_matches = sum(len(rows) for rows in matches.values())
    table = Table(
        title=f"🔍  Results for '{query}'  ({total_matches} matches)",
        box=box.ROUNDED,
        border_style="bright_cyan",
        header_style="bold bright_cyan",
        show_lines=False,
        padding=(0, 2),
    )
    table.add_column("#", style="dim", width=6, justify="right")
    table.add_column("Sheet", style="bright_magenta", width=15)
    table.add_column("Row", style="dim", width=6, justify="right")
    table.add_column("Keyword", style="white", min_width=30)

    match_idx = 1
    for sheet_name in sheet_names:
        if sheet_name not in matches:
            continue
        ws = wb[sheet_name]
        for row_idx in matches[sheet_name]:
            keyword = str(ws.cell(row=row_idx, column=1).value or "")
            lower_kw = keyword.lower()
            start = lower_kw.find(query_lower)
            if start != -1:
                before = keyword[:start]
                match = keyword[start : start + len(query_lower)]
                after = keyword[start + len(query_lower) :]
                display = (
                    f"{before}[bold bright_magenta]{match}[/bold bright_magenta]{after}"
                )
            else:
                display = keyword
            table.add_row(str(match_idx), sheet_name, str(row_idx), display)
            match_idx += 1

    console.print(table)
    console.print()

    # Ask what to do with the matches
    action = questionary.select(
        "What would you like to do with these results?",
        choices=[
            "🟢  Mark all as RELEVANT (green)",
            "🔴  Mark all as IRRELEVANT (red)",
            "⊘  Mark all as NEUTRAL (unmarked)",
            "↩  Back to menu (no changes)",
        ],
        style=custom_style,
    ).ask()

    if action is None or "Back" in action:
        return

    if "NEUTRAL" in action:
        fill = WHITE_FILL
        label = "neutral"
        target_color = None
    elif "IRRELEVANT" in action:
        fill = RED_FILL
        label = "irrelevant"
        target_color = "red"
    else:
        fill = GREEN_FILL
        label = "relevant"
        target_color = "green"

    # Apply color to all matched rows across all sheets (with protection logic)
    marked_count = 0
    skipped_count = 0

    for sheet_name in matches:
        ws = wb[sheet_name]
        for row_idx in matches[sheet_name]:
            row = [ws.cell(row=row_idx, column=1)]
            current_color = get_row_color(row)

            # Protection logic
            if target_color == "red" and current_color == "green":
                skipped_count += 1
                continue
            elif target_color == "green" and current_color == "red":
                skipped_count += 1
                continue

            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill
            marked_count += 1

    # Sort each sheet
    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        sort_sheet(ws)

    # Save
    wb.save(filepath)

    if skipped_count > 0:
        console.print(
            f"\n[bold green]  ✔ Marked {marked_count} rows as {label} and re-sorted all sheets.\n"
            f"  ⊘ Skipped {skipped_count} rows (protected by existing marking)[/bold green]\n"
        )
    else:
        console.print(
            f"\n[bold green]  ✔ Marked {marked_count} rows as {label} and re-sorted all sheets.[/bold green]\n"
        )
    questionary.press_any_key_to_continue(style=custom_style).ask()


def export_clean_multi(wb, sheet_names, filepath):
    """Export cleaned data from multiple sheets to a single new sheet."""
    banner()

    new_name = f"merged_clean"
    counter = 1
    while new_name in wb.sheetnames:
        counter += 1
        new_name = f"merged_clean_{counter}"

    name = questionary.text(
        "Name for the merged clean sheet:",
        default=new_name,
        style=custom_style,
    ).ask()

    if name is None:
        return

    name = name.strip() or new_name

    # Create new sheet
    new_ws = wb.create_sheet(title=name)

    # Add header
    new_ws["A1"] = "Keyword"

    # Copy non-red rows from all sheets
    dest_row = 2
    total_copied = 0
    total_skipped = 0

    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        data_rows = get_data_rows(ws)
        for row_idx in data_rows:
            row_cells = [
                ws.cell(row=row_idx, column=c) for c in range(1, ws.max_column + 1)
            ]
            color = get_row_color(row_cells)
            if color == "red":
                total_skipped += 1
                continue
            for col_idx in range(1, ws.max_column + 1):
                src = ws.cell(row=row_idx, column=col_idx)
                dst = new_ws.cell(row=dest_row, column=col_idx)
                dst.value = src.value
                dst.fill = copy.copy(src.fill)
                dst.font = copy.copy(src.font)
            dest_row += 1
            total_copied += 1

    wb.save(filepath)

    console.print(
        Panel(
            f"[bold green]✔  Created sheet '[bright_cyan]{name}[/bright_cyan]'\n\n"
            f"   Sheets merged:   [white]{len(sheet_names)}[/white]\n"
            f"   Rows copied:     [white]{total_copied}[/white]\n"
            f"   Rows removed:    [red]{total_skipped}[/red] (irrelevant)[/bold green]",
            border_style="green",
            box=box.ROUNDED,
            padding=(1, 3),
        )
    )
    console.print()
    questionary.press_any_key_to_continue(style=custom_style).ask()


def show_instructions():
    """Display help and instructions for using the tool."""
    banner()
    instructions = """
[bold bright_cyan]━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━[/bold bright_cyan]
[bold white]KEYWORD HELPER - USER GUIDE[/bold white]
[bold bright_cyan]━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━[/bold bright_cyan]

[bold green]📁 GETTING STARTED[/bold green]
1. Select an Excel file from the [bright_cyan]input/[/bright_cyan] directory
2. Choose to work with a single sheet or multiple sheets (merged)
3. Use the main menu to perform operations

[bold green]👁 VIEW DATA[/bold green]
• See all keywords in the current sheet(s)
• Shows keyword status: [green]Relevant[/green], [red]Irrelevant[/red], or Unmarked
• Displays statistics: total, relevant, irrelevant counts
• If >500 keywords, data is paginated (500 per page)

[bold green]🔍 KEYWORD SEARCH[/bold green]
• Search for keywords containing a specific substring
• Search is case-insensitive
• View all matching keywords in a results table
• Then choose to:
  → [green]Mark as RELEVANT[/green] (green highlight)
  → [red]Mark as IRRELEVANT[/red] (red highlight)
  → [cyan]Mark as NEUTRAL[/cyan] (remove any marking)
  → Cancel with no changes

[bold green]⚠️ PROTECTION RULES[/bold green]
• If a keyword is already marked [green]Relevant[/green], you cannot mark it [red]Irrelevant[/red]
• If a keyword is already marked [red]Irrelevant[/red], you cannot mark it [green]Relevant[/green]
• Neutral markings can override anything

[bold green]📤 EXPORT CLEAN SHEET[/bold green]
• Creates a new sheet combining only relevant/unmarked keywords
• Removes all [red]Irrelevant[/red] rows automatically
• Choose a custom name for the new sheet
• Works across merged sheets too!

[bold green]🧹 REMOVE DUPLICATE KEYWORDS[/bold green]
• Finds duplicate keywords (case-insensitive)
• Keeps the first occurrence, removes duplicates
• Only works on single sheets
• Warns you before deleting

[bold green]🔗 MULTI-SHEET MODE[/bold green]
• Select multiple sheets when prompted
• All operations work on merged data
• Marks apply to matching keywords across all sheets
• Export creates a single merged clean sheet

[bold green]📄 SWITCH WORKSHEETS[/bold green]
• Change to a different sheet or sheet combination anytime
• Re-run the sheet selection prompt

[bold green]📂 SWITCH FILE[/bold green]
• Load a different Excel file
• Back to sheet selection

[bold green]💾 AUTO-SAVE[/bold green]
• All changes are saved immediately to the Excel file
• Markings, sorts, and deduplication all auto-save

[bold green]📊 AUTO-SORT[/bold green]
• After marking keywords, the sheet auto-sorts:
  1. [green]Relevant[/green] keywords at the top
  2. Unmarked in the middle
  3. [red]Irrelevant[/red] at the bottom

[bold bright_cyan]━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━[/bold bright_cyan]
"""
    console.print(instructions)
    console.print()
    questionary.press_any_key_to_continue(style=custom_style).ask()


# ══════════════════════════════════════════════════════════════════════════════
#  Main loop
# ══════════════════════════════════════════════════════════════════════════════


def main():
    try:
        # Step 1: Pick file
        filepath = pick_file()
        wb = load_workbook(filepath)

        # Step 2: Pick sheet(s)
        sheet_names = pick_sheet(wb)

        # Step 3: Main menu loop
        while True:
            banner()

            # Show brief status
            merged_data = get_multi_sheet_data(wb, sheet_names)
            total = sum(len(rows) for rows in merged_data.values())
            green_count = 0
            red_count = 0
            for sheet_name in sheet_names:
                ws = wb[sheet_name]
                for row_idx, _ in merged_data[sheet_name]:
                    row = [ws.cell(row=row_idx, column=1)]
                    if get_row_color(row) == "green":
                        green_count += 1
                    elif get_row_color(row) == "red":
                        red_count += 1

            sheet_display = (
                sheet_names[0]
                if len(sheet_names) == 1
                else f"{len(sheet_names)} sheets"
            )
            merge_indicator = "" if len(sheet_names) == 1 else " [MERGED]"

            console.print(
                f"  📂 [bright_cyan]{filepath.name}[/bright_cyan]  →  "
                f"📄 [bright_cyan]{sheet_display}[/bright_cyan]{merge_indicator}  |  "
                f"[white]{total}[/white] rows  "
                f"[green]{green_count} ✔[/green]  [red]{red_count} ✘[/red]  "
                f"[dim]{total - green_count - red_count} unmarked[/dim]\n"
            )

            action = questionary.select(
                "What would you like to do?",
                choices=[
                    "👁  View Data",
                    "🔍  Keyword Search",
                    "📤  Export Clean Sheet (remove irrelevant)",
                    "🧹  Remove Duplicate Keywords",
                    "📄  Switch Worksheet(s)",
                    "📂  Switch File",
                    "❓  Help & Instructions",
                    "🚪  Exit",
                ],
                style=custom_style,
                instruction="(↑/↓ to move, Enter to select)",
            ).ask()

            if action is None or "Exit" in action:
                banner()
                console.print("  [dim]Goodbye! 👋[/dim]\n")
                break

            elif "View" in action:
                if len(sheet_names) == 1:
                    view_data(wb[sheet_names[0]])
                else:
                    view_multi_sheets(wb, sheet_names)

            elif "Search" in action:
                if len(sheet_names) == 1:
                    keyword_search(wb[sheet_names[0]], wb, filepath)
                else:
                    keyword_search_multi(wb, sheet_names, filepath)

            elif "Export" in action:
                if len(sheet_names) == 1:
                    export_clean_sheet(wb[sheet_names[0]], wb, filepath)
                else:
                    export_clean_multi(wb, sheet_names, filepath)

            elif "Duplicate" in action:
                if len(sheet_names) == 1:
                    deduplicate_keywords(wb[sheet_names[0]], wb, filepath)
                else:
                    banner()
                    console.print(
                        "[yellow]  Deduplication is per-sheet. Please switch to a single sheet.[/yellow]\n"
                    )
                    questionary.press_any_key_to_continue(style=custom_style).ask()

            elif "Help" in action:
                show_instructions()

            elif "Switch Worksheet" in action:
                sheet_names = pick_sheet(wb)

            elif "Switch File" in action:
                filepath = pick_file()
                wb = load_workbook(filepath)
                sheet_names = pick_sheet(wb)

    except KeyboardInterrupt:
        console.print("\n  [dim]Interrupted. Goodbye! 👋[/dim]\n")
        sys.exit(0)


if __name__ == "__main__":
    main()
